# -------------------------------------------------------------------
# 1. Import 모듈
# -------------------------------------------------------------------
# Python 표준 라이브러리
from datetime import datetime
from dateutil.relativedelta import relativedelta
import random

# Django 및 서드파티 라이브러리
from django.http import HttpResponse
from django.contrib.auth import authenticate
from django.contrib.auth.models import User, Group
from django.db import models, transaction
from django.db.models import Count, Q, Sum, Exists, OuterRef
from django.db.models.functions import Coalesce
from django.utils import timezone
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from rest_framework import viewsets, generics, status
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# 로컬 앱 모듈
from .models import (
    ClientData, EmployeeProfile, Incentive, PerformanceRecord, SiteConfiguration,
    AttendanceRecord
)
from .permissions import IsAdminUser
from .serializers import (
    ClientDataSerializer, IncentiveSerializer, PerformanceRecordSerializer,
    SiteConfigurationSerializer, StaffSerializer, UserSerializer,
    AttendanceRecordSerializer, UserManagementSerializer
)
from .pagination import FiftyResultsSetPagination


# -------------------------------------------------------------------
# 2. 인증 및 사용자 관리 API
# -------------------------------------------------------------------
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class StaffListView(generics.ListAPIView):
    """ 'Staff' 그룹에 속한 모든 사용자의 목록과 출근 상태를 반환합니다. """
    serializer_class = StaffSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    filter_backends = [SearchFilter]
    search_fields = ['username', 'first_name']
    pagination_class = None

    def get_queryset(self):
        today = timezone.now().date()
        # 오늘 출근 기록이 있는지 여부를 'checked_in'이라는 필드로 추가합니다.
        checked_in_today = AttendanceRecord.objects.filter(
            employee=OuterRef('pk'),
            work_date=today
        )
        return User.objects.filter(groups__name='Staff').annotate(
            checked_in=Exists(checked_in_today)
        ).order_by('-checked_in', 'first_name') # 출근한 사람 우선 정렬

@api_view(['POST'])
def login_view(request):
    username = request.data.get('username')
    password = request.data.get('password')
    user = authenticate(username=username, password=password)
    if user is not None:
        token, _ = Token.objects.get_or_create(user=user)
        groups = [group.name for group in user.groups.all()]
        return Response({
            'token': token.key, 'user_id': user.id, 'username': user.username, 'groups': groups,
        }, status=status.HTTP_200_OK)
    return Response({'error': '잘못된 사용자명 또는 비밀번호입니다.'}, status=status.HTTP_401_UNAUTHORIZED)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    request.user.auth_token.delete()
    return Response({'message': '로그아웃 성공'}, status=status.HTTP_200_OK)


# -------------------------------------------------------------------
# 3. 핵심 데이터 관리 API (ViewSet)
# -------------------------------------------------------------------
class ClientDataViewSet(viewsets.ModelViewSet):
    serializer_class = ClientDataSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name']
    ordering_fields = [
        'created_at', 'is_distributed', 'owner__first_name', 'distribution_date',
        'contact', 'address', 'name', 'gender', 'status', 'transmission_status'
    ]
    ordering = ['-created_at']
    pagination_class = FiftyResultsSetPagination

    def get_queryset(self):
        user = self.request.user
        queryset = ClientData.objects.all()
        
        start_date_str = self.request.query_params.get('start_date')
        end_date_str = self.request.query_params.get('end_date')
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__range=[start_date, end_date])
            except (ValueError, TypeError):
                pass

        if self.request.query_params.get('distributed') == 'false':
            queryset = queryset.filter(is_distributed=False)
        
        if user.groups.filter(name='Admin').exists():
            return queryset
        return queryset.filter(owner=user)

class PerformanceRecordViewSet(viewsets.ModelViewSet):
    queryset = PerformanceRecord.objects.all().order_by('-date')
    serializer_class = PerformanceRecordSerializer
    permission_classes = [IsAuthenticated]
    def list(self, request, *args, **kwargs):
        if request.query_params.get('ranking', 'false').lower() == 'true':
            record_type = request.query_params.get('type')
            if not record_type:
                return Response({"error": "랭킹 조회를 위한 'type'이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)
            queryset = User.objects.annotate(
                total_value=Coalesce(Sum('performancerecord__value', filter=Q(performancerecord__record_type=record_type)), 0)
            ).order_by('-total_value')[:3]
            ranked_data = [{'rank': rank, 'employee_username': user.username, 'total_value': user.total_value} 
                           for rank, user in enumerate(queryset, 1)]
            return Response(ranked_data)
        return super().list(request, *args, **kwargs)


# -------------------------------------------------------------------
# 4. 특정 액션 처리 API
# -------------------------------------------------------------------
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@transaction.atomic
def distribute_clients(request):
    """ 고객을 상담사에게 순차 또는 랜덤으로 배분합니다. """
    client_ids = request.data.get('client_ids', [])
    staff_ids = request.data.get('staff_ids', [])
    distribution_date = request.data.get('distribution_date')
    is_random = request.data.get('randomize', False) # 랜덤 배분 여부

    if not all([client_ids, staff_ids, distribution_date]):
        return Response({'error': '고객, 상담사, 배분날짜를 모두 선택해야 합니다.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        staff_users = list(User.objects.filter(id__in=staff_ids, groups__name='Staff'))
        clients_to_distribute = ClientData.objects.filter(id__in=client_ids)
        
        if not staff_users:
            return Response({'error': '유효한 상담사가 없습니다.'}, status=status.HTTP_400_BAD_REQUEST)

        # 랜덤 배분일 경우 상담사 목록을 섞습니다.
        if is_random:
            random.shuffle(staff_users)
        
        staff_count = len(staff_users)
        for i, client in enumerate(clients_to_distribute):
            client.owner = staff_users[i % staff_count]
            client.is_distributed = True
            client.distribution_date = distribution_date
            client.save()

        return Response({'message': f'{len(clients_to_distribute)}명의 고객을 {staff_count}명의 상담사에게 배분했습니다.'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def client_excel_upload(request):
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return Response({'error': '엑셀 파일이 필요합니다.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]: continue
            name, contact, address, note = row[:4]
            ClientData.objects.create(name=name, contact=contact, address=address, note=note)
        return Response({'message': '엑셀 파일이 성공적으로 업로드되었습니다.'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': f'파일 처리 중 오류 발생: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# -------------------------------------------------------------------
# 5. 통계 및 대시보드 API
# -------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_my_summary(request):
    user = request.user
    today = datetime.now()
    user_clients = ClientData.objects.filter(owner=user, created_at__year=today.year, created_at__month=today.month)
    status_counts = user_clients.values('status').annotate(count=Count('status'))
    summary = {'total': user_clients.count(), 'PENDING': 0, 'ABSENT': 0, 'FAIL': 0, 'SUCCESS_1': 0, 'SUCCESS_2': 0, 'PROMISING': 0}
    for item in status_counts:
        if item['status'] in summary:
            summary[item['status']] = item['count']
    success_count = summary['SUCCESS_1'] + summary['SUCCESS_2']
    summary['success_rate'] = (success_count / summary['total'] * 100) if summary['total'] > 0 else 0
    return Response(summary)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def get_performance_statistics(request):
    today = datetime.now()
    SUCCESS_STATUSES = ['SUCCESS_1', 'SUCCESS_2']
    summary_stats = {'total_clients': ClientData.objects.count(),'unassigned_clients': ClientData.objects.filter(is_distributed=False).count(),'total_contracts': ClientData.objects.filter(status__in=SUCCESS_STATUSES).count(),}
    monthly_stats = {'new_clients': ClientData.objects.filter(created_at__year=today.year, created_at__month=today.month).count(),'contracts': ClientData.objects.filter(created_at__year=today.year, created_at__month=today.month, status__in=SUCCESS_STATUSES).count(),}
    region_top5 = list(ClientData.objects.values('address').annotate(count=Count('id')).order_by('-count')[:5])
    monthly_trend = []
    for i in range(6):
        date_cursor = today - relativedelta(months=i)
        contracts_in_month = ClientData.objects.filter(created_at__year=date_cursor.year, created_at__month=date_cursor.month, status__in=SUCCESS_STATUSES).count()
        monthly_trend.append({'month': date_cursor.strftime("%Y-%m"), 'contracts': contracts_in_month})
    monthly_trend.reverse()
    return Response({'summary': summary_stats,'monthly_performance': monthly_stats,'region_top5': region_top5,'monthly_contract_trend': monthly_trend,}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_incentive_board_data(request):
    today = datetime.now()
    SUCCESS_STATUSES = ['SUCCESS_1', 'SUCCESS_2']
    incentive_rules = list(Incentive.objects.order_by('case_count').values('case_count', 'reward_amount'))
    staff_users = User.objects.filter(groups__name='Staff')
    board_data = []
    for user in staff_users:
        success_count = ClientData.objects.filter(owner=user, status__in=SUCCESS_STATUSES, updated_at__year=today.year, updated_at__month=today.month).count()
        current_reward = 0
        for rule in incentive_rules:
            try:
                if str(rule['case_count']).isdigit() and success_count >= int(rule['case_count']):
                    current_reward = rule['reward_amount']
                elif '~' in str(rule['case_count']):
                    start, end = map(int, str(rule['case_count']).split('~'))
                    if start <= success_count <= end:
                        current_reward = rule['reward_amount']
            except (ValueError, TypeError):
                continue
        board_data.append({'employee_name': user.first_name or user.username, 'success_count': success_count, 'reward_amount': current_reward})
    sorted_board_data = sorted(board_data, key=lambda x: x['success_count'], reverse=True)
    return Response(sorted_board_data, status=status.HTTP_200_OK)


# -------------------------------------------------------------------
# 6. 기타 설정 API (ViewSet)
# -------------------------------------------------------------------
class IncentiveViewSet(viewsets.ModelViewSet):
    queryset = Incentive.objects.all().order_by('case_count')
    serializer_class = IncentiveSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            self.permission_classes = [IsAuthenticated]
        else:
            self.permission_classes = [IsAuthenticated, IsAdminUser]
        return super().get_permissions()
    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update(self, request):
        Incentive.objects.all().delete()
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class SiteConfigurationViewSet(viewsets.ModelViewSet):
    queryset = SiteConfiguration.objects.all()
    serializer_class = SiteConfigurationSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = None
    lookup_field = 'key'


# -------------------------------------------------------------------
# 7. 출퇴근 기록 관리 API
# -------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_today_attendance_status(request):
    record = AttendanceRecord.objects.filter(employee=request.user, work_date=timezone.now().date()).first()
    if record:
        serializer = AttendanceRecordSerializer(record)
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response({}, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def check_in_view(request):
    employee = request.user
    today = timezone.now().date()
    if AttendanceRecord.objects.filter(employee=employee, work_date=today).exists():
        return Response({'error': '이미 오늘 출근 처리되었습니다.'}, status=status.HTTP_400_BAD_REQUEST)
    record = AttendanceRecord.objects.create(employee=employee)
    serializer = AttendanceRecordSerializer(record)
    return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def check_out_view(request):
    employee = request.user
    today = timezone.now().date()
    try:
        record = AttendanceRecord.objects.get(employee=employee, work_date=today)
    except AttendanceRecord.DoesNotExist:
        return Response({'error': '출근 기록이 없습니다. 출근 먼저 해주세요.'}, status=status.HTTP_404_NOT_FOUND)
    if record.check_out_time:
        return Response({'error': '이미 퇴근 처리되었습니다.'}, status=status.HTTP_400_BAD_REQUEST)
    record.check_out_time = timezone.now()
    record.save()
    serializer = AttendanceRecordSerializer(record)
    return Response(serializer.data, status=status.HTTP_200_OK)

class AttendanceRecordListView(generics.ListAPIView):
    serializer_class = AttendanceRecordSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = None
    def get_queryset(self):
        month_str = self.request.query_params.get('month')
        if not month_str:
            today = timezone.now()
            return AttendanceRecord.objects.filter(work_date__year=today.year, work_date__month=today.month)
        try:
            year, month = map(int, month_str.split('-'))
            return AttendanceRecord.objects.filter(work_date__year=year, work_date__month=month)
        except (ValueError, TypeError):
            return AttendanceRecord.objects.none()


# -------------------------------------------------------------------
# 8. 관리자용 직원 관리 API
# -------------------------------------------------------------------
class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserManagementSerializer
    permission_classes = [IsAuthenticated, IsAdminUser]
    pagination_class = None
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        is_active = request.data.get('is_active')
        if is_active is not None:
            instance.is_active = is_active
        group_names = request.data.get('groups')
        if group_names is not None:
            instance.groups.clear()
            for group_name in group_names:
                try:
                    group = Group.objects.get(name=group_name)
                    instance.groups.add(group)
                except Group.DoesNotExist:
                    pass
        instance.save()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


# -------------------------------------------------------------------
# 9. 엑셀 다운로드 API
# -------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def download_clients_excel(request):
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    queryset = ClientData.objects.all().order_by('-created_at')
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__range=[start_date, end_date])
            filename = f"client_data_{start_date_str}_to_{end_date_str}.xlsx"
        except (ValueError, TypeError):
            filename = "client_data_all.xlsx"
    else:
        filename = "client_data_all.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '고객 데이터'
    headers = ['고객명', '연락처', '주소', '상담사', '가입일', '상태', '메모']
    worksheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for client in queryset:
        consultant_name = client.consultant if hasattr(client, 'consultant') else (client.owner.first_name if client.owner else '미지정')
        row = [
            client.name, client.contact, client.address, consultant_name,
            client.created_at.strftime('%Y-%m-%d %H:%M'),
            client.get_status_display(), client.note
        ]
        worksheet.append(row)
    workbook.save(response)
    return response
